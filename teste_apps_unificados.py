import streamlit as st
import pandas as pd
import requests
import json
import datetime
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime as dt
from PIL import Image

# ========================= CONFIGURAÇÃO GERAL =========================
st.set_page_config(
    page_title="Operações - Lojas MIMI",
    page_icon="🧠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ========================= ESTILO =========================
st.markdown("""
    <style>
    .stButton > button {
        background-color: #0047AB;
        color: white;
        font-weight: bold;
    }
    .stDownloadButton > button {
        background-color: #28A745;
        color: white;
        font-weight: bold;
    }
    .stSelectbox, .stTextInput, .stNumberInput {
        border: 1px solid #ccc;
        border-radius: 5px;
    }
    .big-font {
        font-size: 22px !important;
    }
    .small-font {
        font-size: 14px;
        color: gray;
    }
    </style>
""", unsafe_allow_html=True)

# ========================= MENU LATERAL =========================
logo = Image.open("logo_lojas_mimi.jpeg")
st.sidebar.image(logo, use_container_width=True)
st.sidebar.markdown("## 📁 Menu de Operações")
menu = st.sidebar.radio(
    "Escolha a operação:",
    [
        "♻️ Processo de Trocas",
        "🛍️ Processo de Pedidos",
        "📦 Transferência entre Lojas",
        "🔍 Pesquisa de Produtos",
        "🛠️ Atualizador de Preços",
        "🔎 Procura de Fornecedor"  
    ]
)
st.title("🧠 Sistema de Operações - Lojas MIMI")

# ========================= FUNÇÕES COMUNS =========================
@st.cache_data(show_spinner=False)
def carregar_csv_combinado():
    url = "https://raw.githubusercontent.com/LojasMimi/transferencia_loja/refs/heads/main/cad_concatenado.csv"
    df = pd.read_csv(url, dtype=str).fillna("")
    df = df.loc[:, ~df.columns.str.contains("^Unnamed", case=False)]
    df.columns = df.columns.str.strip().str.upper()
    def dedup_columns(cols):
        seen = {}
        new = []
        for c in cols:
            if c in seen:
                seen[c] += 1
                new.append(f"{c}_{seen[c]}")
            else:
                seen[c] = 0
                new.append(c)
        return new
    df.columns = dedup_columns(df.columns)
    if "SITUACAO" in df.columns:
        df["SITUACAO"] = df["SITUACAO"].str.replace("ç", "c", regex=False)
    if "DESCRIÇÃO" in df.columns:
        df["DESCRIÇÃO"] = df["DESCRIÇÃO"].str.replace("ç", "c", regex=False)
    return df

def buscar_produto(codigo, coluna, df):
    codigo = str(codigo).strip()
    resultado = df[df[coluna].astype(str).str.strip() == codigo]
    return resultado.iloc[0] if not resultado.empty else None

# ========================= APP 1: TROCAS =========================
def app_trocas():
    st.header("♻️ Processo de Trocas")
    st.divider()

    if "trocas_dados" not in st.session_state:
        st.session_state.trocas_dados = []

    df = carregar_csv_combinado()
    fornecedores = sorted(df["FORNECEDOR"].dropna().unique())

    aba1, aba2 = st.tabs(["🧍 Troca Individual", "📂 Troca por Lote"])

    # ========================= ABA 1 - INDIVIDUAL =========================
    with aba1:
        sel = st.selectbox("Fornecedor:", [""] + fornecedores)
        if sel:
            df_f = df[df["FORNECEDOR"] == sel]
            st.subheader("🔍 Buscar Produto para Troca")
            c1, c2, c3 = st.columns([3, 4, 2])
            tipo = c1.selectbox("Buscar por:", ["CÓDIGO DE BARRAS", "REF"])
            col = "CODIGO BARRA" if tipo == "CÓDIGO DE BARRAS" else "CODIGO"
            ids = sorted(df_f[col].dropna().astype(str).str.strip().unique())
            ident = c2.selectbox(tipo + ":", [""] + ids)
            qtd = c3.number_input("Quantidade", 1, step=1, value=1)

            if st.button("🔎 Buscar Produto para Troca"):
                if not ident:
                    st.warning("Selecione um identificador válido.")
                else:
                    res = buscar_produto(ident, col, df_f)
                    if res is not None:
                        st.session_state.trocas_dados.append({
                            "CODIGO BARRA": res.get("CODIGO BARRA", ""),
                            "CODIGO": res.get("CODIGO", ""),
                            "FORNECEDOR": res.get("FORNECEDOR", ""),
                            "DESCRICAO": res.get("DESCRICAO", ""),
                            "QUANTIDADE": qtd
                        })
                        st.success(f"Adicionado: {res.get('DESCRICAO', '')}")
                    else:
                        st.warning("Produto não encontrado.")
        else:
            st.info("Selecione um fornecedor.")

    # ========================= ABA 2 - LOTE =========================
    with aba2:
        c1, c2 = st.columns(2)

        if c1.button("📥 Baixar Modelo Excel"):
            modelo = pd.DataFrame(columns=["CODIGO BARRA", "CODIGO", "QTD"])
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                modelo.to_excel(writer, index=False, sheet_name="Trocas")
            buf.seek(0)
            st.download_button("⬇️ Baixar Modelo", buf, "modelo_troca.xlsx")

        fornecedor_lote = c2.selectbox("Fornecedor para Lote:", fornecedores)
        tipo_id = st.selectbox("Usar como identificador:", ["CÓDIGO DE BARRAS", "REF"])
        col_id = "CODIGO BARRA" if tipo_id == "CÓDIGO DE BARRAS" else "CODIGO"
        arquivo = st.file_uploader("📤 Enviar Excel com Trocas", type=["xlsx"])

        if arquivo and fornecedor_lote:
            try:
                df_up = pd.read_excel(arquivo).fillna("")
                if not all(c in df_up.columns for c in ["CODIGO BARRA", "CODIGO", "QTD"]):
                    st.error("Arquivo inválido. Verifique se as colunas estão corretas: CODIGO BARRA, CODIGO, QTD.")
                else:
                    df_forn = df[df["FORNECEDOR"] == fornecedor_lote]
                    faltando_qtd = False
                    adicionados = 0

                    for _, row in df_up.iterrows():
                        ident = str(row.get(col_id, "")).strip()
                        qtd_raw = str(row.get("QTD", "")).strip()

                        if not qtd_raw.isdigit():
                            faltando_qtd = True
                            continue

                        qtd = int(qtd_raw)
                        res = buscar_produto(ident, col_id, df_forn)
                        if res is not None:
                            st.session_state.trocas_dados.append({
                                "CODIGO BARRA": res.get("CODIGO BARRA", ""),
                                "CODIGO": res.get("CODIGO", ""),
                                "FORNECEDOR": res.get("FORNECEDOR", ""),
                                "DESCRICAO": res.get("DESCRICAO", ""),
                                "QUANTIDADE": qtd
                            })
                            adicionados += 1

                    if faltando_qtd:
                        st.warning("⚠️ Há produtos com a QTD vazia ou inválida. Por favor, preencha todas as linhas corretamente.")
                    if adicionados:
                        st.success(f"✅ {adicionados} produtos adicionados com sucesso.")
            except Exception as e:
                st.error(f"Erro ao processar o arquivo: {e}")

    # ========================= RESULTADOS GERAIS =========================
    if st.session_state.trocas_dados:
        df_t = pd.DataFrame(st.session_state.trocas_dados)
        st.subheader(f"📋 Itens adicionados ({len(df_t)}):")
        st.dataframe(df_t, use_container_width=True)
        cA, cB = st.columns([1, 3])

        if cA.button("🗑️ Remover Último"):
            rem = st.session_state.trocas_dados.pop()
            st.warning(f"Removido: {rem['DESCRICAO']}")

        def ger_excel(dados):
            provs = set(i['FORNECEDOR'] for i in dados)
            if len(provs) > 1:
                return None, "Múltiplos fornecedores."
            try:
                wb = load_workbook("FORM-TROCAS.xlsx")
                ws = wb.active
                ws["C3"] = provs.pop()
                for i, item in enumerate(dados[:27]):
                    r = 6 + i
                    ws[f"A{r}"] = item["CODIGO BARRA"]
                    ws[f"B{r}"] = item["CODIGO"]
                    ws[f"C{r}"] = item["DESCRICAO"]
                    ws[f"D{r}"] = item["QUANTIDADE"]
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf, None
            except Exception as e:
                return None, str(e)

        if cB.button("📄 Gerar Formulário"):
            ex, err = ger_excel(st.session_state.trocas_dados)
            if err:
                st.error(err)
            else:
                st.success("Formulário pronto!")
                st.download_button("📥 Baixar", ex, "FORMULARIO_TROCA.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Nenhum item adicionado.")


# ========================= APP 2: PEDIDOS =========================
def app_pedidos():
    st.header("🛍️ Processo de Pedidos")
    st.divider()
    
    if "produtos_solicitados" not in st.session_state:
        st.session_state.produtos_solicitados = []

    df = carregar_csv_combinado()
    aba1, aba2, aba3 = st.tabs(["🧍 Individual", "📂 Lote", "📋 Revisão"])

    # --- Aba 1: Individual ---
    with aba1:
        forn = st.selectbox("Fornecedor:", sorted(df["FORNECEDOR"].dropna().unique()))
        tipo = st.selectbox("Buscar por:", ["CÓDIGO DE BARRAS", "REF"])
        col = "CODIGO BARRA" if tipo == "CÓDIGO DE BARRAS" else "CODIGO"
        df_f = df[df["FORNECEDOR"] == forn]
        opc = st.selectbox("Produto:", sorted(df_f[col].dropna().unique()))
        qtd = st.number_input("Quantidade:", 1, step=1)

        if st.button("➕ Adicionar Pedido"):
            prod = df_f[df_f[col] == opc]
            if not prod.empty:
                p = prod.iloc[0]
                it = {
                    "FORNECEDOR": forn,
                    "CODIGO BARRA": p["CODIGO BARRA"],
                    "CODIGO": p["CODIGO"],
                    "DESCRICAO": p["DESCRICAO"],
                    "QTD": qtd,
                    "ORIGEM": p.get("__ORIGEM_PLANILHA__", "")
                }
                st.session_state.produtos_solicitados.append(it)
                st.toast("✅ Produto adicionado!")
            else:
                st.error("Produto não encontrado.")

    # --- Aba 2: Lote ---
    with aba2:
        c1, c2 = st.columns(2)
        if c1.button("📥 Baixar Modelo Excel"):
            modelo = pd.DataFrame(columns=["CODIGO BARRA", "CODIGO", "QTD"])
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as w:
                modelo.to_excel(w, index=False, sheet_name="Modelo")
            buf.seek(0)
            st.download_button("⬇️", buf, "modelo_pedido.xlsx")

        fornecedor_lote = c2.selectbox("Fornecedor para Lote:", sorted(df["FORNECEDOR"].dropna().unique()))
        arq = c2.file_uploader("📤 Enviar Excel", type=["xlsx"])
        tipo_col = st.selectbox("Usar como identificador:", ["CÓDIGO DE BARRAS", "REF"])
        col_id = "CODIGO BARRA" if tipo_col == "CÓDIGO DE BARRAS" else "CODIGO"

        if arq:
            wb = load_workbook(filename=BytesIO(arq.read()))
            ws = wb.active
            data = ws.values
            cols = next(data)
            df_l = pd.DataFrame(data, columns=cols).fillna("")

            # Filtro apenas produtos do fornecedor selecionado
            df_forn = df[df["FORNECEDOR"] == fornecedor_lote]
            qtd_faltante = False

            for _, row in df_l.iterrows():
                cod = str(row.get(col_id, "")).strip()
                qtd_raw = str(row.get("QTD", "")).strip()

                # Verifica se quantidade está presente e válida
                if not qtd_raw.isdigit():
                    qtd_faltante = True
                    continue  # pula o item com quantidade inválida

                qtd = int(qtd_raw)
                prod = df_forn[df_forn[col_id] == cod]

                if not prod.empty:
                    p = prod.iloc[0]
                    it = {
                        "FORNECEDOR": p["FORNECEDOR"],
                        "CODIGO BARRA": p["CODIGO BARRA"],
                        "CODIGO": p["CODIGO"],
                        "DESCRICAO": p["DESCRICAO"],
                        "QTD": qtd,
                        "ORIGEM": p.get("__ORIGEM_PLANILHA__", "")
                    }
                    st.session_state.produtos_solicitados.append(it)

            if qtd_faltante:
                st.warning("⚠️ Há valores faltantes ou inválidos na coluna QTD. Os itens com erro foram ignorados.")
            else:
                st.toast("✅ Produtos adicionados!")

    # --- Aba 3: Revisão ---
    with aba3:
        if st.session_state.produtos_solicitados:
            df_f = pd.DataFrame(st.session_state.produtos_solicitados)
            st.dataframe(df_f, use_container_width=True)

            if st.button("📤 Gerar Planilha Final"):
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine='openpyxl') as w:
                    df_f.to_excel(w, index=False, sheet_name="Pedidos")
                buf.seek(0)
                st.success("Planilha pronta!")
                st.download_button("⬇️", buf, "pedidos.xlsx")
        else:
            st.info("Nenhum pedido adicionado.")

# ========================= APP 3: TRANSFERÊNCIAS =========================
def app_transferencias():
    st.header("📦 Transferência entre Lojas")
    st.divider()
    if "formulario_dados" not in st.session_state:
        st.session_state.formulario_dados=[]
    lojas = ["MIMI","KAMI","TOTAL MIX","E-COMMERCE"]
    c1,c2 = st.columns(2)
    de_loja = c1.selectbox("Loja de Origem", lojas)
    para_loja = c2.selectbox("Loja de Destino", [l for l in lojas if l!=de_loja])
    df = carregar_csv_combinado()
    modo = st.radio("Modo:", ["Individual","Lote"],horizontal=True)
    if modo == "Lote":
        if st.button("📥 Baixar Modelo Excel"):
            modelo = pd.DataFrame(columns=["CODIGO BARRA", "QUANTIDADE"])
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                modelo.to_excel(writer, index=False, sheet_name="Transferencia")
            buf.seek(0)
            st.download_button("⬇️ Baixar Modelo", data=buf, file_name="modelo_transferencia.xlsx")

        up = st.file_uploader("📤 Upload Planilha", type=["xlsx"])
        if up:
            df_l = pd.read_excel(up)
            for _, row in df_l.iterrows():
                cod = str(row["CODIGO BARRA"]).strip()
                qtd = int(row["QUANTIDADE"])
                prod = buscar_produto(cod, "CODIGO BARRA", df)
                if prod is not None:
                    st.session_state.formulario_dados.append({
                        "CODIGO BARRA": cod,
                        "CODIGO": prod.get("CODIGO", ""),
                        "FORNECEDOR": prod.get("FORNECEDOR", ""),
                        "DESCRICAO": prod.get("DESCRICAO", ""),
                        "QUANTIDADE": qtd
                    })

    else:
        t,v,q = st.columns([2,3,2])
        tipo = t.selectbox("Buscar por:", ["Código de Barras","REF"])
        val = v.text_input("Valor:")
        qtd = q.number_input("Quantidade",1,step=1,value=1)
        if st.button("➕ Adicionar Produto"):
            col = "CODIGO BARRA" if tipo=="Código de Barras" else "CODIGO"
            prod = buscar_produto(val, col, df)
            if prod is not None:
                st.session_state.formulario_dados.append({
                    "CODIGO BARRA": prod["CODIGO BARRA"],
                    "CODIGO": prod["CODIGO"],
                    "FORNECEDOR": prod["FORNECEDOR"],
                    "DESCRICAO": prod["DESCRICAO"],
                    "QUANTIDADE": qtd
                })
                st.toast("✅ Produto adicionado!")
    if st.session_state.formulario_dados:
        df_f = pd.DataFrame(st.session_state.formulario_dados)
        st.dataframe(df_f, use_container_width=True)
        if st.button("📄 Gerar Relatório"):
            wb = load_workbook("FORMULÁRIO DE TRANSFERENCIA ENTRE LOJAS.xlsx")
            ws = wb.active
            ws["A4"] = f"DE: {de_loja}"
            ws["C4"] = para_loja
            ws["D4"] = "DATA " + dt.today().strftime("%d/%m/%Y")
            for i,item in enumerate(st.session_state.formulario_dados[:30]):
                r = 8+i
                ws[f"A{r}"]=item["CODIGO BARRA"]
                ws[f"B{r}"]=item["CODIGO"]
                ws[f"C{r}"]=item["FORNECEDOR"]
                ws[f"D{r}"]=item["DESCRICAO"]
                ws[f"E{r}"]=item["QUANTIDADE"]
            buf=BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("⬇️ Baixar", buf, "TRANSFERENCIA.xlsx")

# ========================= APP 4: PESQUISA DE PRODUTOS (API) =========================
def app_pesquisa():
    st.header("🔍 Pesquisa de Produtos (API Varejo Fácil)")
    st.divider()
    st.markdown("<p class='small-font'>Consulta em tempo real na base do Varejo Fácil</p>", unsafe_allow_html=True)
    cod = st.text_input("📦 Código de barras", placeholder="Ex: 7891234567890")
    if st.button("🔎 Consultar Produto"):
        if not cod.strip():
            st.warning("Digite um código de barras válido.")
        else:
            url1 = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/consulta/0{cod}"
            hdr = {'x-api-key': st.secrets.api.x_api_key, 'Cookie': st.secrets.api.cookie}
            r1 = requests.get(url1, headers=hdr)
            if r1.status_code==200:
                dp = r1.json()
                if 'id' in dp and 'descricao' in dp:
                    pid = dp['id']; desc = dp['descricao']
                    st.success("✅ Produto encontrado!")
                    st.markdown(f"<div class='big-font'><strong>📄 Descrição:</strong> {desc}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div class='small-font'>🆔 ID: {pid}</div>", unsafe_allow_html=True)
                    r2 = requests.get(f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/{pid}/precos", headers=hdr)
                    if r2.status_code==200:
                        lp = r2.json()
                        p1 = next((i for i in lp if i.get("lojaId")==1), None)
                        if p1:
                            v = p1.get("precoVenda1","N/A"); c = p1.get("custoProduto","N/A")
                            with st.expander("💰 Preço e Custo"):
                                st.write(f"**Preço de Venda:** R$ {v:.2f}" if isinstance(v,(int,float)) else f"**Preço de Venda:** {v}")
                                st.write(f"**Custo:** R$ {c:.2f}" if isinstance(c,(int,float)) else f"**Custo:** {c}")
                        else:
                            st.info("Sem dados de preço para esta loja.")
                    else:
                        st.error(f"Erro ao consultar preços: {r2.status_code}")
                else:
                    st.warning("Produto não encontrado ou dados incompletos.")
            else:
                st.error(f"Erro ao buscar produto: {r1.status_code}")

# ========================= APP 5: ATUALIZADOR DE PREÇOS =========================
def app_atualizador_precos():
    st.header("🛠️ Atualizador de Preços")
    st.markdown("Atualize preço de Venda ou Custo via API Varejo Fácil")
    def fmt(c): return c.zfill(13) if len(c)<13 else c
    def login(u,p):
        r = requests.post("https://lojasmimi.varejofacil.com/api/auth",
            headers={"Content-Type":"application/json"},
            data=json.dumps({"username":u,"password":p}))
        if r.status_code==200: return r.json().get("accessToken")
    def obter_id(cb, tok):
        r = requests.get(f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/consulta/0{cb}",
            headers={"Authorization":tok})
        if r.status_code==200:
            d=r.json(); return d.get("id"),d.get("descricao")
    def obter_custos(pid, tok):
        r = requests.get(f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/{pid}/precos",
            headers={"Authorization":tok})
        if r.status_code==200: return r.json()
    def atualiza(custos, novo, tipo, tok):
        data=dt.now().astimezone().isoformat(); ok=[]
        for c in custos:
            if c['lojaId'] in [1,2,5]:
                pld={k:c.get(k) for k in ["id","lojaId","produtoId",
                    "precoVenda1","custoProduto","precoMedioDeReposicao","precoFiscalDeReposicao"]}
                pld["dataUltimoReajustePreco1"]=data
                if tipo=="Venda": pld["precoVenda1"]=novo
                else:
                    pld["custoProduto"]=novo
                    pld["precoMedioDeReposicao"]=novo
                    pld["precoFiscalDeReposicao"]=novo
                r = requests.put(f"https://lojasmimi.varejofacil.com/api/v1/produto/precos/{c['id']}",
                    headers={"Content-Type":"application/json","Authorization":tok},
                    data=json.dumps(pld))
                if r.status_code==200: ok.append(c['lojaId'])
        return ok

    if "access_token" not in st.session_state:
        st.subheader("🔐 Login")
        u = st.text_input("Usuário")
        p = st.text_input("Senha", type="password")
        if st.button("Entrar"):
            with st.spinner("Validando..."):
                t=login(u,p)
            if t: st.session_state.update(access_token=t, usuario=u); st.success("✅ Logado!"); st.rerun()
            else: st.error("Credenciais inválidas.")
    else:
        st.success(f"Usuário: {st.session_state.usuario}")
        if st.button("🚪 Sair"):
            st.session_state.pop("access_token", None); st.session_state.pop("usuario", None); st.rerun()
        st.divider()
        col1, col2 = st.columns(2)
        metodo = col1.selectbox("Buscar por", ["Código de Barras","ProdutoId"])
        tipo = col2.selectbox("Tipo de atualização", ["Venda","Custo"])
        entr = st.text_input(f"Insira {metodo}")
        if entr:
            if metodo=="Código de Barras":
                cb = fmt(entr)
                pid, desc = obter_id(cb, st.session_state.access_token)
            else:
                try:
                    pid = int(entr); desc=f"Produto ID {pid}"
                except:
                    st.error("ID inválido."); return
            if pid:
                st.write(f"**Produto:** {desc}")
                custos = obter_custos(pid, st.session_state.access_token)
                if custos:
                    dfc = pd.DataFrame([{"Loja":c['lojaId'], "Preço Venda":c.get("precoVenda1",0),
                        "Custo":c.get("custoProduto",0)} for c in custos if c['lojaId'] in [1,2,5]])
                    st.dataframe(dfc, use_container_width=True)
                    novo = st.number_input("Novo valor (R$)", min_value=0.0, step=0.01)
                    if st.button("Atualizar Preço"):
                        ok = atualiza(custos, novo, tipo, st.session_state.access_token)
                        if ok: st.success(f"✅ Atualizado em lojas: {', '.join(map(str, ok))}")
                        else: st.warning("Nenhuma loja atualizada.")
                else:
                    st.error("Não foi possível obter preços.")
            else:
                st.error("Produto não encontrado.")

# ========================= APP 6: PROCURA DE FORNECEDOR =========================

def app_procura_fornecedor():
    st.header("🔎 Procura de Fornecedor")
    st.divider()
    
    df = carregar_csv_combinado()
    
    if "__ORIGEM_PLANILHA__" not in df.columns:
        st.error("A coluna '__ORIGEM_PLANILHA__' não foi encontrada no dataset.")
        return

    fornecedores = sorted(df["FORNECEDOR"].dropna().unique())
    selecionados = st.multiselect("Selecione os fornecedores que deseja localizar:", fornecedores)

    if selecionados:
        resultado = (
            df[df["FORNECEDOR"].isin(selecionados)][["FORNECEDOR", "__ORIGEM_PLANILHA__"]]
            .drop_duplicates()
            .sort_values(by="FORNECEDOR")
            .rename(columns={"__ORIGEM_PLANILHA__": "PLANILHA DE ORIGEM"})
        )

        st.subheader(f"📍 Origem dos fornecedores selecionados ({len(resultado)})")
        st.dataframe(resultado, use_container_width=True)
    else:
        st.info("Selecione ao menos um fornecedor para visualizar as origens.")


# ========================= ROTEAMENTO =========================
if menu == "♻️ Processo de Trocas":
    app_trocas()
elif menu == "🛍️ Processo de Pedidos":
    app_pedidos()
elif menu == "📦 Transferência entre Lojas":
    app_transferencias()
elif menu == "🔍 Pesquisa de Produtos":
    app_pesquisa()
elif menu == "🛠️ Atualizador de Preços":
    app_atualizador_precos()
elif menu == "🔎 Procura de Fornecedor":
    app_procura_fornecedor()

# ========================= RODAPÉ =========================
st.markdown("""
<hr style="margin-top: 40px; margin-bottom: 10px;">
<div style='text-align: center; font-size: 13px; color: gray;'>
Desenvolvido por <strong>Pablo</strong> · Lojas MIMI © 2025
</div>
""", unsafe_allow_html=True)
